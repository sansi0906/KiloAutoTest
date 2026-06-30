import base64
import os
import re
import time
import glob
import random
import sys
import cv2
import numpy as np
import easyocr
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from recognize_all import (
    correct_char, correct_text, smart_complete,
    convert_expression, calculate_expression, recognize_with_multiple_params
)

PROVINCES = ["110000","120000","130000","140000","150000","210000","220000","230000","310000","320000","330000","340000","350000","360000","370000","410000","420000","430000","440000","450000","460000","500000","510000","520000","530000","540000","610000","620000","630000","640000","650000","660000"]

CAPTCHA_URL = 'https://dljz.mof.gov.cn/captchaImage'
QUERY_URL = 'https://dljz.mof.gov.cn/guild/guild/getHomeList'

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
IMG_DIR = os.path.join(BASE_DIR, 'img')
EXCEL_FILE = os.path.join(BASE_DIR, 'guild_data.xlsx')

reader = None

def init_ocr():
    global reader
    reader = easyocr.Reader(['ch_sim', 'en'], gpu=False)
    print("EasyOCR模型加载完成")

def get_captcha(max_retries=50, retry_delay=600):
    os.makedirs(IMG_DIR, exist_ok=True)
    
    for attempt in range(1, max_retries + 1):
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'application/json, text/javascript, */*; q=0.01',
                'Accept-Language': 'zh-CN,zh;q=0.9',
            }
            response = requests.get(CAPTCHA_URL, headers=headers, timeout=15)
            data = response.json()
            
            uuid = data.get('uuid', '')
            img_data = data.get('img', '')
            
            if not uuid or not img_data:
                print(f'[{attempt}] 缺少uuid或图片数据，等待{retry_delay//60}分钟后重试...')
                time.sleep(retry_delay)
                continue
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(IMG_DIR, f"captcha_{timestamp}.jpg")
            
            with open(filename, 'wb') as f:
                f.write(base64.b64decode(img_data))
            
            corrected = recognize_with_multiple_params(filename, reader)
            
            print(f'[{attempt}] UUID: {uuid}, 识别: {corrected}')
            
            if not corrected:
                print(f'[{attempt}] 未识别到内容，等待{retry_delay//60}分钟后重试...')
                time.sleep(retry_delay)
                continue
            
            expr_result, expr = calculate_expression(corrected)
            if expr_result is None:
                print(f'[{attempt}] 计算失败: {expr}，等待{retry_delay//60}分钟后重试...')
                time.sleep(retry_delay)
                continue
            
            print(f'[{attempt}] 计算成功: {corrected} = {expr_result}')
            return uuid, int(expr_result)
            
        except Exception as e:
            print(f'[{attempt}] 请求失败: {e}，等待{retry_delay//60}分钟后重试...')
            time.sleep(retry_delay)
    
    raise Exception(f'验证码获取失败，已达 {max_retries} 次最大重试')

def query_data(uuid, ynum, province):
    params = {
        'auditLevel': '',
        'orgType': '',
        'guildName': '',
        'guildStatus': '',
        'divisionProvince': province,
        'pageNum': 1,
        'pageSize': 20,
        'code': ynum,
        'codeUuid': uuid,
    }
    
    all_rows = []
    page_num = 1
    
    while True:
        params['pageNum'] = page_num
        try:
            response = requests.get(QUERY_URL, params=params, headers={'User-Agent': 'Mozilla/5.0'}, timeout=15)
            data = response.json()
            
            if data.get('code') == 400 or '验证码' in (data.get('msg', '') or ''):
                return all_rows if all_rows else None, '验证码错误'
            
            rows = data.get('rows', [])
            if not rows:
                break
            
            all_rows.extend(rows)
            print(f'  第{page_num}页: {len(rows)}条数据')
            page_num += 1
            time.sleep(1)
            
        except Exception as e:
            print(f'  请求第{page_num}页失败: {e}')
            return all_rows if all_rows else None, '请求失败'
    
    return all_rows, None

def save_to_excel(rows):
    headers = ['guildName', 'regisCno', 'registerAddress', 'divisionName', 'guildStatus']
    
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
    
    existing_keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = (row[0], row[1], row[2], row[3], row[4] if len(row) > 4 else '')
        existing_keys.add(key)
    
    new_count = 0
    for row in rows:
        data = [
            row.get('guildName', ''),
            row.get('regisCno', ''),
            row.get('registerAddress', ''),
            row.get('divisionName', ''),
            row.get('guildStatus', ''),
        ]
        key = tuple(data)
        if key not in existing_keys:
            ws.append(data)
            existing_keys.add(key)
            new_count += 1
    
    wb.save(EXCEL_FILE)
    if new_count > 0:
        print(f'  已保存 {new_count} 条新数据到 {EXCEL_FILE}')
    else:
        print(f'  没有新数据，跳过保存')


def deduplicate_excel():
    if not os.path.exists(EXCEL_FILE):
        return
    
    headers = ['guildName', 'regisCno', 'registerAddress', 'divisionName', 'guildStatus']
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    rows = [headers]
    seen = set()
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = (row[0], row[1], row[2], row[3], row[4] if len(row) > 4 else '')
        if key not in seen:
            seen.add(key)
            rows.append(row)
    
    wb2 = Workbook()
    ws2 = wb2.active
    for row in rows:
        ws2.append(row)
    
    wb2.save(EXCEL_FILE)
    print(f'  去重完成，保留 {len(rows)-1} 条数据')

def main():
    print('='*60)
    print('代理记账行业协会数据爬虫 (优化版)')
    print(f'图片目录: {IMG_DIR}')
    print(f'数据文件: {EXCEL_FILE}')
    print('='*60)
    
    init_ocr()
    try:
        deduplicate_excel()
    except PermissionError:
        print('  Excel文件被占用，跳过启动时去重')
    
    round_num = 1
    while True:
        print(f'\n--- 第 {round_num} 轮开始 ---')
        
        for province in PROVINCES:
            print(f'\n处理省份: {province}')
            
            # 每个省份都先获取验证码
            while True:
                try:
                    uuid, ynum = get_captcha()
                    break  # 验证码获取成功，跳出重试循环
                except Exception as e:
                    print(f'  获取验证码失败: {e}')
                    print('  等待10分钟后重试...')
                    time.sleep(600)
                    continue  # 继续重试当前省份
            
            print(f'  获取成功 - UUID: {uuid}, code: {ynum}')
            
            while True:
                rows, error = query_data(uuid, ynum, province)
                
                if error == '验证码错误':
                    if rows:
                        save_to_excel(rows)
                        print(f'  省份 {province} 部分数据已保存（验证码失效），共 {len(rows)} 条')
                    print(f'  验证码错误，等待10分钟后重新获取...')
                    time.sleep(600)
                    try:
                        uuid, ynum = get_captcha(retry_delay=5)
                        print(f'  新验证码 - UUID: {uuid}, code: {ynum}')
                        continue
                    except Exception as e:
                        print(f'  重新获取验证码失败: {e}')
                        continue  # 继续重试当前省份
                
                if error == '请求失败':
                    if rows:
                        save_to_excel(rows)
                        print(f'  省份 {province} 部分数据已保存（请求失败），共 {len(rows)} 条')
                    print(f'  请求失败，等待10分钟后重新获取验证码...')
                    time.sleep(600)
                    try:
                        uuid, ynum = get_captcha(retry_delay=5)
                        print(f'  新验证码 - UUID: {uuid}, code: {ynum}')
                        continue
                    except Exception as e:
                        print(f'  重新获取验证码失败: {e}')
                        continue  # 继续重试当前省份
                
                if rows:
                    save_to_excel(rows)
                    print(f'  省份 {province} 处理完成，共 {len(rows)} 条')
                
                break
            
            print(f'  省份 {province} 已完成，继续下一个省份')
        
        print(f'\n第 {round_num} 轮结束')
        print(f'等待10分钟后开始下一轮...')
        
        for i in range(600, 0, -1):
            if i % 60 == 0:
                print(f'  剩余 {i//60} 分钟')
            time.sleep(1)
        
        round_num += 1

if __name__ == '__main__':
    main()
